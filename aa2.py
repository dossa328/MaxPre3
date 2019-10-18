import openpyxl

wb = openpyxl.load_workbook('result2.xlsx')

sheet = wb['Sheet1']
# sheet['A1'] = 'hello'
sheet.cell(row=1, column=1, value='되나1')
sheet.cell(row=1, column=2, value='되나2')
sheet.cell(row=1, column=3, value='되나3')
sheet.cell(row=3, column=3, value='되나4')

print(sheet.cell(row=1, column=1).value)
print("%s" % sheet.max_column)

wb.save('result2.xlsx')
