import json
import random
import threading
import alpha_pruning
import alpha_pruning2
import straight_forward
import straight_forward_for_datasetting
import dijkstra_distributer

import time
import openpyxl
from Metro import Metro

with open('trans.json', 'r', encoding='UTF-8') as json_file:
    trans_data = json.load(json_file)

with open('line.json', 'r', encoding='UTF-8') as line_file:
    line_data = json.load(line_file)

# with open('trans_classification_experiment_data.json', 'r', encoding='UTF-8') as data_file:
with open('trans_classification_list.json', 'r', encoding='UTF-8') as data_file:
    input_data = json.load(data_file)

data_set = []

for i in line_data:
    for j in line_data[i]:
        data_set.append(j + i)


def rand_start_end(_data_set):
    while True:
        _start = random.choice(_data_set)
        _end = random.choice(_data_set)

        if _start is not _end:
            break

    return _start, _end


alpha_for_alpha_pruning = 1.2
alpha_for_straight_forward = 1.0
metro = Metro()
alpha_pruning_result = []
alpha_pruning_result_avg = 0
straight_forward_result = []
straight_forward_result_avg = 0

wb = openpyxl.load_workbook('result for cal.xlsx')
sheet = wb['Sheet1']

# for k in range(1, 11):
# start_time = time.time()
reresult = []
path_start_end_pairs = {}
compare_paths_transdata = []


a = 0
for j in range(0, len(input_data)):
    print(a)
    # reresult.extend(dijkstra_distributer.cal_dists(input_data[str(j)]['from'], input_data[str(j)]['to']))
    # reresult.extend(alpha_pruning.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_alpha_pruning))
    # reresult.append(alpha_pruning2.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_alpha_pruning)[0])
    reresult.append(alpha_pruning2.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_alpha_pruning)[0])
    a = a + 1


# set_of_reresult = []
# for d in range(len(reresult)):
#     set_of_reresult.append(reresult[0])


index = 1
for v in reresult:
    compare_paths_transdata = []
    # compare_paths_transdata = list(set(v).intersection(trans_data["trans"]))
    for v2 in trans_data["trans"]:
        if v2 in v[0]:
            compare_paths_transdata.append(v2)
    print(compare_paths_transdata)
    len_compare_paths_transdata = len(compare_paths_transdata)
    # if v[0] in compare_paths_transdata and v[0] in trans_data["trans"]:
    #     len_compare_paths_transdata = len_compare_paths_transdata - 1
    # if v[-1] in compare_paths_transdata and v[-1] in trans_data["trans"]:
    #     len_compare_paths_transdata = len_compare_paths_transdata - 1

    path_start_end_pairs[len_compare_paths_transdata] = (v[0], v[-1])ㅔ
    print(len_compare_paths_transdata, ",", v[0], ",", v[-1])
    print(v[0][0], v[0][-1])
    print(v)
    # 지나가는 경유역 개수 (동일역 포함
    sheet.cell(row=index, column=1, value=len_compare_paths_transdata)
    sheet.cell(row=index, column=2, value=v[0][0])
    sheet.cell(row=index, column=3, value=v[0][-1])
    sheet.cell(row=index, column=4, value=v[-1])
    sheet.cell(row=index, column=5, value=str(v[0]))


    # sheet.cell(row=index, column=4, value=v)
    # sheet.cell(row=index, column=4).value = v
    # for r in range(0, len(v)):
    #     sheet.cell(row=index, column=4).value=v[r]
    index = index + 1

    # alpha_pruning_result.append(alpha_pruning.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_alpha_pruning))
    # print(alpha_pruning.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_alpha_pruning))
    # alpha_pruning_result_avg = alpha_pruning_result_avg + alpha_pruning_result[j][2]

wb.save('result for cal.xlsx')


# end_time = time.time()
# print(alpha_pruning_result_avg / len(input_data))
# print("AP_WorkingTime: {} sec".format(end_time-start_time))
    #
    # sheet.cell(row=k, column=1, value="알파 프루닝 결과 평균치")
    # sheet.cell(row=k, column=2, value=alpha_pruning_result_avg / len(input_data))
    # sheet.cell(row=k, column=3, value="알파 프루닝 결과 시간")
    # sheet.cell(row=k, column=4, value=format(end_time-start_time))
    #
    # wb.save('result.xlsx')
# start_time = time.time()
# for j in range(len(input_data)):
#     straight_forward_result.append(straight_forward.get_result(metro, input_data[str(j)]['from'], input_data[str(j)]['to'], alpha_for_straight_forward))
#     straight_forward_result_avg = straight_forward_result_avg + straight_forward_result[j][2]
#
# end_time = time.time()
# print(straight_forward_result_avg / len(input_data))
# print("SF_WorkingTime: {} sec".format(end_time - start_time))



# # 1.2 내 포함되는 데이터셋 만드는 코드
# time_out = 3
# deny = []
# for i in range(0, 10000):
#     inlist = []
#     ran_start, ran_end = rand_start_end(data_set)
#
#     done_counting = threading.Event()
#     # t = threading.Thread(target=alpha_pruning.get_result, args=(metro, ran_start, ran_end, alpha))
#     # t = threading.Thread(target=straight_forward.get_result, args=(metro, ran_start, ran_end, alpha_for_alpha_pruning))
#     t = threading.Thread(target=straight_forward_for_datasetting.get_result, args=(metro, ran_start, ran_end, alpha_for_alpha_pruning))
#     t.setDaemon(True)
#
#     t.start()
#
#     t.join(time_out)
#     done_counting.wait(time_out)
#     # if runtime out (difficult problem)
#     if t.is_alive():
#         # print(ran_start, ran_end)
#         pass
#     # if runtime out (easy problem)
#     else:
#         inlist = [ran_start, ran_end]
#         if not list in deny:
#             deny.append([ran_start, ran_end])
#             # sheet.cell(row=i+1, column=1, value=ran_start)
#             # sheet.cell(row=i+1, column=2, value=ran_end)
#             print(ran_start, ran_end)
#             # wb.save('result2.xlsx')
#         else:
#             i = i - 1
#     # pass

# wb.save('result2.xlsx')
